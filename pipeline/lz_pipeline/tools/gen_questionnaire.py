"""Generate the customer-facing LZ Assessment Questionnaire FROM the schema.

Pre-engagement discovery instrument: open-ended questions a customer answers
alone, before any workshop. A filled copy is converted to a draft
lz.spec.<customer>.json by the /questionnaire-to-spec skill (via
tools/dump_questionnaire.py) — gaps become the decisions-needed list.

The question CATALOGUE below is the source of truth; the xlsx is a rendering.
Each question carries wiring: the "Sheet.Table[.field]" targets it informs.
A coverage check walks lz_spec/schema.py and fails the build if any
non-exempt table is unreachable from every question — so schema drift
surfaces here, not in a stale questionnaire.

Usage: py tools/gen_questionnaire.py [-o out.xlsx] [--check]
"""

import argparse
import datetime
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent   # workspace
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "lz_spec"))

QUESTIONNAIRE_VERSION = "1.1"
DEFAULT_OUT = ROOT / "HuaweiCloud-LZ-Assessment-Questionnaire.xlsx"

# Tables no customer question should cover: engineer-derived plumbing, or
# reserved/not-implemented. Everything else in schema.SHEETS must be wired.
ENGINEER_ONLY = {
    "01_Foundation.EnabledPolicyTypes",   # fixed org policy-type enum
    "01_Foundation.TrustedServices",      # baseline org integrations
    "03_Identity.ServiceAgencies",        # service-to-service trust plumbing
    "05_Network.ERAvailabilityZones",     # derived from region
    "05_Network.HubERAttachments",        # derived from hub topology
    "05_Network.SpokeERAttachments",      # derived from spoke topology
    "05_Network.RAMSharePrincipals",      # derived from account list
    "06_Observability.LogConverge",       # schema says: leave empty, derived
    "11_SGACL.NetworkACLs",               # RESERVED (LZR-031)
    "11_SGACL.ACLRules",                  # RESERVED (LZR-031)
}

CATEGORIES = [
    "Organization & Accounts",
    "Identity & Access",
    "Network",
    "Security",
    "Compliance & Audit",
    "Operations & Monitoring",
    "Finance & Cost Management",
]

# (tier, category, question, guidance, wiring, default_if_silent)
#   tier: "core" | "deep"
#   wiring: list of "Sheet.Table" or "Sheet.Table.field" targets (may be
#           empty for pure-context questions)
Q = []


def _q(tier, cat, question, guidance, wiring, default="", example=""):
    assert cat in CATEGORIES, cat
    Q.append({"tier": tier, "category": cat, "question": question,
              "guidance": guidance, "wiring": wiring, "default": default,
              "example": example})


# ── Organization & Accounts ─────────────────────────────────────────────────
_q("core", "Organization & Accounts",
   "What does your cloud environment look like today? Please list the cloud providers you use, roughly how many accounts, subscriptions, or projects you have, and what they are used for.",
   "A simple summary is enough. If you already have an account inventory or organization diagram, attach it instead of retyping the details.",
   ["01_Foundation.OrganizationalUnits"],
   example="AWS: ~25 accounts, one per team; Azure: 8 subscriptions (legacy ERP); no Huawei footprint yet.")
_q("core", "Organization & Accounts",
   "Do you already use Huawei Cloud? If yes, what is running there today, and are any existing accounts expected to remain in the future landing zone?",
   "We can design around existing accounts where needed. A new account structure is usually simpler to govern, but migration constraints may require a mixed approach.",
   ["01_Foundation.Settings", "01_Foundation.CoreAccounts"],
   "Greenfield: a new organization with fresh core + workload accounts.")
_q("core", "Organization & Accounts",
   "Which applications or workloads are in the first migration or deployment wave? For each one, tell us which environments you run, such as dev, test, UAT, and production.",
   "You can use Appendix A if a table is easier. Also tell us whether you prefer separate accounts for each environment or another isolation model.",
   ["01_Foundation.WorkloadAccounts", "01_Foundation.OrganizationalUnits"],
   "One account per workload per environment tier.")
_q("core", "Organization & Accounts",
   "Do you already have a landing zone or cloud account hierarchy elsewhere, such as Azure management groups or AWS Organizations? What would you like to keep or improve?",
   "A diagram or export is useful. We can reuse governance patterns that already work well for your teams.",
   ["01_Foundation.OrganizationalUnits"],
   example="AWS Control Tower: Security / Infrastructure / Workloads OUs. Worked well; too many SCP exceptions crept in.")
_q("core", "Organization & Accounts",
   "Which Huawei Cloud region will be your primary region? Do you need a secondary region for disaster recovery, and are there any data residency requirements?",
   "For example, Singapore may be the primary region with another approved region reserved for DR. Residency requirements may also affect organization-wide controls.",
   ["Global.Settings.home_region", "03_Identity.RegisteredRegions",
    "04_Perimeter.SCPs.AllowedRegions"],
   "Single region ap-southeast-3, region-lock guardrail staged (not enforced).")
_q("core", "Organization & Accounts",
   "What email address pattern should Huawei Cloud account mailboxes use? Each account needs a globally unique email address, for example cloud-<account>@yourco.com.",
   "Plus-addressing, such as team+lz-prod@yourco.com, works if your mail system supports it. If you do not have a pattern yet, we can propose one for review.",
   ["01_Foundation.CoreAccounts.Email", "01_Foundation.WorkloadAccounts.Email"],
   "cloud-<account>@<your-domain> pattern proposed in the draft.",
   example="cloud-<account>@acme.com, e.g. cloud-lz-prod@acme.com (plus-addressing not supported).")
_q("core", "Organization & Accounts",
   "Do you have a naming convention for cloud resources such as VPCs, subnets, gateways, buckets, and log groups? If yes, please share the pattern and an example, such as <org>-<region>-<env>-<service>-<nn> for acme-sg-prd-vpc-01.",
   "One answer lets us name every resource in the draft design consistently. Please note any hard limits, such as bucket names being globally unique and lowercase.",
   ["05_Network.HubVPCs.VPCName", "05_Network.HubSubnets.Name",
    "05_Network.SpokeVPCs.VPCName", "05_Network.EIPs.Name",
    "05_Network.NATGateways.Name", "06_Observability.AuditSettings",
    "08_DNS.ResolverEndpoints.Name"],
   "Huawei best-practice naming (<org>-<region>-<env>-<service>-<nn>); all resource names inferred from it in the draft spec, flagged for review.",
   example="<org>-<region>-<env>-<service>-<nn>, e.g. acme-sg-prd-vpc-01; buckets lowercase with acme- prefix.")
_q("deep", "Organization & Accounts",
   "How do you expect your Huawei Cloud account estate to grow over the next 12 to 24 months? Who should approve new accounts?",
   "Consider new workloads, teams, business units, countries, acquisitions, or temporary project accounts.",
   ["01_Foundation.OrganizationalUnits", "01_Foundation.WorkloadAccounts"])

# ── Identity & Access ───────────────────────────────────────────────────────
_q("core", "Identity & Access",
   "Which teams will use or manage Huawei Cloud, and what is each team responsible for?",
   "Typical teams include platform, application, security, network, database, finance, and audit. Appendix C can be used if you prefer a table.",
   ["03_Identity.Groups"],
   example="Platform team (LZ + network), AppDev (portal), SecOps (SOC + audit), DBA; finance needs read-only cost views.")
_q("core", "Identity & Access",
   "How do you want users to sign in to Huawei Cloud? For example, through Microsoft Entra ID, AD FS, Okta, another identity provider, or Huawei Cloud-managed users.",
   "If you use an external identity provider, please include the product and version where known.",
   ["03_Identity.Settings", "03_Identity.Users"],
   "Native Identity Center users until IdP federation is confirmed.")
_q("core", "Identity & Access",
   "If you plan to federate sign-in, does your identity provider support SAML 2.0 and SCIM provisioning?",
   "SAML provides federated sign-in. SCIM can automate user and group provisioning. If you are unsure, your identity team can confirm this later.",
   ["03_Identity.Settings"],
   example="Entra ID - SAML 2.0 and SCIM both available.")
_q("core", "Identity & Access",
   "Will vendors, MSPs, auditors, or other third parties need access to Huawei Cloud? If yes, what should they access and for how long?",
   "Please include the expected permission level and whether access should be permanent or time-limited.",
   ["03_Identity.PermissionSets", "03_Identity.AccountAssignments"],
   "No third-party access provisioned.")
_q("core", "Identity & Access",
   "What identity security policies must Huawei Cloud follow for passwords, MFA, account lockout, and session timeouts?",
   "If your organization already has an identity or access policy, attach it. If not, we can propose a baseline for review rather than asking you to define every setting now.",
   ["03_Identity.Settings.ic_min_password_length", "03_Identity.Settings.ic_mfa_required"],
   "LZ baseline hardening policy.")
_q("deep", "Identity & Access",
   "What CI/CD and automation tools do you use, and where do they run? How should pipelines authenticate to Huawei Cloud?",
   "Examples include GitHub Actions, GitLab, Jenkins, and Azure DevOps. Where possible, we prefer short-lived federated credentials over long-lived access keys.",
   ["03_Identity.PermissionSets", "03_Identity.AppPermissionSets"],
   example="GitHub Actions (cloud-hosted); OIDC federation preferred, no static keys in pipelines.")
_q("deep", "Identity & Access",
   "How do you handle emergency or break-glass access today?",
   "Please describe who can use emergency credentials, how MFA is enforced, where credentials are stored, and how emergency access is reviewed or audited.",
   ["03_Identity.Settings"],
   example="Root credentials sealed with IT security, MFA on; every use requires a ticket and is reviewed.")
_q("deep", "Identity & Access",
   "How detailed do access permissions need to be? Are standard admin, power-user, and read-only roles enough, or do some teams need access limited to specific applications or resources?",
   "Also share any preferred session duration or sign-in portal naming requirements.",
   ["03_Identity.PermissionSets", "03_Identity.AppPermissionSets",
    "03_Identity.AccountAssignments", "01_Foundation.Settings.identity_center_alias"],
   "Three standard tiers per account; no app-scoped sets.")

# ── Network ─────────────────────────────────────────────────────────────────
_q("core", "Network",
   "Please describe your current network at a high level. Include data centres, offices, existing cloud networks, and how they connect today.",
   "A topology diagram is preferred if you have one. A rough diagram is also fine at this stage.",
   ["05_Network.HubVPCs"],
   example="Two DCs (SG + KL) linked by MPLS; AWS VPCs reach on-prem via Transit Gateway; offices on SD-WAN.")
_q("core", "Network",
   "How should Huawei Cloud connect to your data centres, offices, or other clouds? Do you expect Direct Connect, site-to-site VPN, or both?",
   "For each connection, note whether high availability is required. Detailed VPN and routing information is covered in the deep-dive section.",
   ["10_VPN.Gateways", "10_VPN.Connections", "10_VPN.Settings"],
   "No hybrid connectivity provisioned.")
_q("core", "Network",
   "How should internet connectivity work for your workloads? Which systems need outbound internet access, and which applications need to be reachable from the internet?",
   "Also tell us whether you are comfortable centralizing internet egress through a shared NAT and firewall. We can review exceptions where needed.",
   ["05_Network.NATGateways", "05_Network.SNATRules", "05_Network.Settings.snat_vpc_attachment"],
   "Centralized egress via hub NAT behind the cloud firewall.")
_q("core", "Network",
   "Do you have any requirements or concerns for connectivity between VPCs, such as bandwidth, latency, traffic inspection, or cost?",
   "A common design uses Enterprise Router for hub-and-spoke connectivity and centralized inspection. We will confirm whether that pattern fits your traffic and cost requirements.",
   ["05_Network.EnterpriseRouter", "05_Network.Settings"],
   "Enterprise Router hub-and-spoke accepted.")
_q("core", "Network",
   "Can you allocate a dedicated private IP range to Huawei Cloud that does not overlap with your existing networks? If yes, what range is available?",
   "Appendix B can be used to list available and reserved ranges. We can either work from a customer-provided subnet plan or propose one for review.",
   ["05_Network.Settings.spoke_private_supernet", "05_Network.HubVPCs",
    "05_Network.HubSubnets", "05_Network.SpokeVPCs", "05_Network.SpokeSubnets"],
   "Huawei plans subnets within a customer-provided supernet.",
   example="10.20.0.0/16 reserved for Huawei Cloud (on-prem uses 10.0-10.19); Huawei plans the subnets.")
_q("core", "Network",
   "How do you expect VPCs to be organized across accounts and applications?",
   "For example, you may use one VPC per workload account, multiple VPCs for separate applications, or another model. Tell us about any isolation requirements that influence this choice.",
   ["05_Network.SpokeVPCs"],
   "One VPC per workload account.")
_q("core", "Network",
   "Which applications or environments need to communicate with each other, and which must stay isolated?",
   "If you have a source, destination, port, and protocol matrix, attach it. Otherwise, a few high-level rules are enough for the first design.",
   ["09_CFW.ACLRules", "09_CFW.AddressGroups", "09_CFW.ServiceGroups",
    "11_SGACL.SecurityGroups", "11_SGACL.SGRules"],
   "Default-deny between applications; shared services reachable by all.")
_q("core", "Network",
   "Which tags should be mandatory on cloud resources, such as application, owner, environment, project, or cost centre?",
   "Also tell us whether missing tags should block deployment or only generate a compliance finding.",
   ["Global.MasterDefaultTags", "01_Foundation.TagPolicies",
    "04_Perimeter.PredefinedTags", "04_Perimeter.SCPs.MandatoryTags",
    "01_Foundation.Settings.enforce_tag_keys_scp"],
   "project/owner/env/bu tagged by the platform; enforcement staged, not blocking.")
_q("deep", "Network",
   "For each site-to-site VPN, please provide the VPN device, public IP, routing method, on-premises networks, expected throughput, and availability requirements.",
   "If you use BGP, include your on-premises ASN. Dual tunnels with dynamic routing are usually preferred for high availability.",
   ["10_VPN.Gateways", "10_VPN.CustomerGateways", "10_VPN.Connections",
    "05_Network.EnterpriseRouter.er_asn"],
   example="HQ: Fortinet 200F, static IP, BGP (AS 65010), reach 10.0.0.0/12, ~200 Mbit/s, dual tunnels.")
_q("deep", "Network",
   "If you plan to use more than one Huawei Cloud region, what traffic needs to move between regions and why?",
   "Examples include data replication, DR failover, backup traffic, or user traffic. Include rough bandwidth needs where known.",
   [],
   example="Skip - single region for now; DR ambition revisited next year.")
_q("deep", "Network",
   "How should subnets be divided inside each VPC?",
   "Tell us whether you separate by application tier, availability zone, or another standard. Include any subnet sizing rules you already follow.",
   ["05_Network.SpokeSubnets", "05_Network.HubSubnets"],
   "app + db subnets per spoke VPC.")
_q("deep", "Network",
   "Do you run or plan to run Kubernetes or containers on Huawei Cloud? If yes, which workloads are involved?",
   "Container platforms can consume significant IP space, so early cluster estimates help us size the network correctly.",
   ["05_Network.SpokeVPCs"],
   example="CCE planned for the portal re-platform next year; reserve a /20 per cluster in the IP plan.")
_q("deep", "Network",
   "How should DNS work between Huawei Cloud and your existing environment?",
   "Please include internal domain names, current DNS servers, required cloud-to-on-premises and on-premises-to-cloud resolution, public DNS hosting needs, and whether DNS query logging is required.",
   ["08_DNS.Settings", "08_DNS.PublicZones", "08_DNS.PrivateZones",
    "08_DNS.RecordSets", "08_DNS.ResolverEndpoints", "08_DNS.ResolverRules",
    "08_DNS.AccessLogs", "05_Network.Settings.subnet_dns_servers"],
   "Private zone + hybrid resolver; public DNS stays at the current registrar.")
_q("deep", "Network",
   "Which applications will be exposed to the internet, and where should TLS terminate?",
   "For each application, note whether TLS should terminate at the load balancer, WAF, or application, and whether you will provide existing certificates.",
   ["05_Network.DNATRules", "05_Network.ELBs"],
   "No inbound publishing provisioned.")
_q("deep", "Network",
   "What internet bandwidth and public IP capacity do you expect to need?",
   "A rough peak outbound bandwidth and estimated number of public IPs is enough. If you have a billing preference, note whether you prefer bandwidth-based or traffic-based charging.",
   ["05_Network.EIPs"],
   "1x 100 Mbit/s egress EIP, bandwidth-billed.")
_q("deep", "Network",
   "Do you need VPC flow logs for troubleshooting, security analysis, or compliance? If yes, how long should they be retained?",
   "Flow logs increase log volume and storage cost, so retention should match the operational or compliance need.",
   ["05_Network.Settings.enable_vpc_flow_logs", "05_Network.Settings.flow_log_retention_days"],
   "Flow logs on, 90-day retention.")

# ── Security ────────────────────────────────────────────────────────────────
_q("core", "Security",
   "Which security tools do you use today, and which of them need to integrate with Huawei Cloud?",
   "Examples include firewalls, endpoint security, SIEM, vulnerability management, privileged access management, and security monitoring platforms.",
   ["09_CFW.Settings"],
   example="Palo Alto perimeter, CrowdStrike EDR, Splunk SIEM - EDR and SIEM must extend to cloud servers.")
_q("core", "Security",
   "What are your expectations for Cloud Firewall? Please cover intrusion prevention, perimeter antivirus if required, alerting, and any preferred billing model.",
   "If you do not have a defined policy yet, we can propose an initial observe-and-tune approach before moving selected protections into blocking mode.",
   ["05_Network.CloudFirewall", "09_CFW.Settings.enable_attack_alarm",
    "09_CFW.Settings.enable_traffic_alarm", "09_CFW.Settings.alarm_topic_name",
    "09_CFW.Settings.enable_anti_virus"],
   "IPS observe mode, alarms on to the ops topic, pay-per-use.")
_q("core", "Security",
   "Which Huawei Cloud security services do you expect to use, if any, such as SecMaster, Host Security Service, or Database Security Service? Who will review the alerts?",
   "It is fine to mark this as undecided. We can recommend services based on your workload and compliance requirements.",
   ["07_Security.Settings", "07_Security.SecMasterModules", "07_Security.AlertRules"],
   "SecMaster in the security account; HSS/DBSS deferred.")
_q("core", "Security",
   "Are there actions that should be blocked across the whole organization, even for administrators?",
   "Examples include disabling audit logging, making storage public, leaving the organization, or deploying resources in unapproved regions.",
   ["04_Perimeter.SCPs"],
   "The 8 LZ baseline guardrails, staged (created, not enforced).")
_q("deep", "Security",
   "Which internet-facing applications need Web Application Firewall protection?",
   "For each application, provide the domain, rough bandwidth or request rate if known, and how TLS certificates are managed.",
   ["07_Security.WAF", "07_Security.WAFDomains"],
   "WAF off until an internet-facing app needs it.")
_q("deep", "Security",
   "Do any public endpoints need specific DDoS thresholds, alerting, or enhanced protection?",
   "If you have experienced DDoS incidents before, include the affected services and any traffic patterns you know.",
   ["07_Security.AntiDDoS"],
   "Default Anti-DDoS thresholds, no per-EIP tuning.")
_q("deep", "Security",
   "Do you maintain IP or domain allowlists, blocklists, threat feeds, or geo-blocking rules that should be carried into Huawei Cloud?",
   "Attach existing lists where possible. We can use them as input to firewall policy design.",
   ["09_CFW.BlackWhiteLists", "09_CFW.DomainGroups"],
   "No seed lists; firewall starts with the rule baseline only.")
_q("deep", "Security",
   "Do you have specific encryption or key-management requirements?",
   "Examples include customer-managed KMS keys, BYOK, HSM-backed keys, key rotation periods, separation of key administrators, and audit requirements.",
   ["06_Observability.AuditSettings.kms_audit_alias"],
   "Platform-managed KMS keys per sensitive bucket.")
_q("deep", "Security",
   "Will any object storage need to be publicly accessible?",
   "If yes, describe the use case and how exceptions should be approved. Otherwise, we can treat public storage as prohibited by default.",
   ["04_Perimeter.SCPs.ExceptionTagKey"],
   "Public storage denied outright, no exception path.")

# ── Compliance & Audit ──────────────────────────────────────────────────────
_q("core", "Compliance & Audit",
   "What are your audit logging requirements? Should activity from all accounts be centralized, protected from changes, and searchable? How long must logs be retained?",
   "Please separate online or searchable retention from long-term archive retention if your policy defines both.",
   ["06_Observability.AuditSettings"],
   "Org-wide audit trail to the security account, 365-day retention.")
_q("core", "Compliance & Audit",
   "Which regulatory, security, or industry standards apply to this environment?",
   "Examples include MAS TRM, PDPA, PCI DSS, ISO 27001, and SOC 2. Also tell us whether auditors need platform-generated evidence or reports.",
   ["04_Perimeter.ConfigConformancePacks"],
   "Landing Zone best-practice pack only.",
   example="ISO 27001 + PDPA; auditors want a yearly conformance evidence export.")
_q("core", "Compliance & Audit",
   "Do you use continuous compliance controls today, such as Azure Policy or AWS Config? Which team should own compliance monitoring and findings in Huawei Cloud?",
   "If you already have a security governance process, describe how cloud findings should feed into it.",
   ["04_Perimeter.ConfigSetup"],
   "Config recorder + org aggregator in the security account.")
_q("deep", "Compliance & Audit",
   "Should any accounts or environments be excluded from standard compliance checks?",
   "Examples might include temporary sandboxes or proof-of-concept accounts. Please explain the reason and approval process for any exemption.",
   ["04_Perimeter.ConfigConformancePacks.ExcludedAccounts"],
   "No exemptions.")

# ── Operations & Monitoring ─────────────────────────────────────────────────
_q("core", "Operations & Monitoring",
   "How do you monitor infrastructure today? Which monitoring tools, metrics, logs, alert rules, and on-call processes should carry over to Huawei Cloud?",
   "We will use this to map existing operational practices to Huawei Cloud monitoring and alerting services.",
   ["06_Observability.OpsSettings", "06_Observability.OneClickNamespaces"],
   "Cloud Eye baseline bundles on core services, alerts to one ops topic.")
_q("core", "Operations & Monitoring",
   "Do you want logs and monitoring data from multiple accounts collected centrally? If yes, which log types are most important?",
   "Common examples include audit, firewall, DNS, VPC flow, operating system, and application logs.",
   ["06_Observability.LogAggregation"],
   "All platform logs converge centrally and archive to OBS.")
_q("core", "Operations & Monitoring",
   "How long should each type of log remain searchable, archived, or moved to lower-cost storage?",
   "If different log classes have different retention periods, list them separately. Approximate targets are fine for the first design.",
   ["06_Observability.LogAggregation.archive_retention_days",
    "06_Observability.AuditSettings.audit_retention_days"],
   "90-day hot, 365-day archive.")
_q("deep", "Operations & Monitoring",
   "Who should receive platform and security alerts, and through which channels?",
   "List the relevant teams, email addresses, or other notification endpoints. You can also indicate different recipients by severity.",
   ["06_Observability.Subscribers", "06_Observability.OpsSettings.topic_name"],
   example="ops-team@acme.com (all severities); oncall SMS +65 9xxx xxxx (critical only).")
_q("deep", "Operations & Monitoring",
   "Do you use a SIEM such as Microsoft Sentinel, Splunk, or QRadar? Which Huawei Cloud logs should be sent to it?",
   "If known, include your preferred ingestion method. We can confirm the technical integration during design.",
   ["06_Observability.LogAggregation.enable_log_aggregation"],
   "No SIEM integration.")
_q("deep", "Operations & Monitoring",
   "What are your backup and disaster recovery requirements?",
   "For each important workload, include what must be protected, target RPO and RTO, retention, and any long-term archive requirements.",
   [],
   "Recorded for the workload phase; no platform default.",
   example="All servers nightly (RPO 24h); databases RPO 15 min; keep 30 days hot + 1 year archive.")

# ── Finance & Cost Management ───────────────────────────────────────────────
_q("core", "Finance & Cost Management",
   "Who owns cloud cost management and billing? Who needs visibility into budgets, forecasts, and account-level spend?",
   "Please include the finance or cloud governance teams that should receive budget notifications or reports.",
   ["02_Finance.Settings"],
   "Central billing from the management account.")
_q("core", "Finance & Cost Management",
   "How should Huawei Cloud costs be allocated internally?",
   "Common dimensions include business unit, department, application, project, environment, and owner. We can align this with your tagging and account structure.",
   ["02_Finance.CostCenters"],
   "Cost centres per business unit, prod/dev typed.")
_q("core", "Finance & Cost Management",
   "How should resources be grouped inside accounts using Enterprise Projects, Huawei Cloud's in-account grouping construct? For example by application, environment, department, or cost centre. Enterprise Projects are similar to Azure resource groups.",
   "Enterprise Projects scope both cost reporting and permissions, so the grouping outlives billing. If you have no preference, we derive a layout from your workload and cost-allocation answers.",
   ["02_Finance.CostCenters", "03_Identity.AppPermissionSets"],
   "One EP per application per environment tier, derived from the workload list; flagged for review.",
   example="Group by app + environment, e.g. portal-prd-ep / portal-uat-ep - mirrors our Azure resource-group layout.")
_q("deep", "Finance & Cost Management",
   "How should shared cloud costs be handled?",
   "For shared network, security, logging, and platform services, tell us whether costs should be absorbed centrally, shown back to teams, or charged back.",
   ["02_Finance.CostCenters"],
   "Platform costs absorbed centrally.")
_q("deep", "Finance & Cost Management",
   "Do you have an existing Huawei Cloud commercial or billing arrangement?",
   "Examples include an enterprise agreement, partner or reseller arrangement, or committed spend. Also include any budget thresholds that should trigger finance alerts. The commercial arrangement itself is engagement context (it stays with the delivery team, not the spec); only the alert recipients and thresholds land in the specification.",
   ["06_Observability.Subscribers"],
   example="Partner/reseller agreement via <partner>; alert finance@acme.com at 80% of monthly budget.")

# ── Appendices (optional structured tables) ─────────────────────────────────
APPENDICES = [
    {
        "ref": "A", "name": "Appendix A - Accounts",
        "title": "Accounts & Environments (optional)",
        "note": "One row per workload per environment. Fill what you know; leave the rest blank.",
        "columns": ["Application / Workload", "Environment", "Proposed account name",
                    "Grouping / OU", "Owner email", "Notes"],
        "example": ["Customer portal", "Production", "app-portal-prod", "Workloads/Prod",
                    "cloud-portal-prod@example.com", "Internet-facing"],
        "wiring": ["01_Foundation.WorkloadAccounts", "01_Foundation.OrganizationalUnits"],
    },
    {
        "ref": "B", "name": "Appendix B - IP Plan",
        "title": "IP Plan (optional)",
        "note": "Known allocations and ranges to avoid. Scope: supernet offered to Huawei Cloud, an existing on-prem/other-cloud range to avoid, or a specific VPC you want pinned.",
        "columns": ["Scope", "CIDR", "Purpose / VPC name", "Account / Environment", "Notes"],
        "example": ["Available supernet", "10.20.0.0/16", "Huawei Cloud allocation", "All",
                    "Nothing else uses this block"],
        "wiring": ["05_Network.Settings.spoke_private_supernet", "05_Network.HubVPCs",
                   "05_Network.HubSubnets", "05_Network.SpokeVPCs", "05_Network.SpokeSubnets"],
    },
    {
        "ref": "C", "name": "Appendix C - Teams",
        "title": "Teams & Access (optional)",
        "note": "One row per team. Access level: admin / power user / read-only / custom (describe in Notes).",
        "columns": ["Team", "Responsibilities", "Members (name + email)",
                    "Access level", "Accounts or applications in scope", "Notes"],
        "example": ["Platform engineering", "Landing zone, network, shared services",
                    "Alice Tan <alice@example.com>", "Administrator", "All accounts", ""],
        "wiring": ["03_Identity.Groups", "03_Identity.Users",
                   "03_Identity.PermissionSets", "03_Identity.AccountAssignments"],
    },
]


# ── Coverage check ──────────────────────────────────────────────────────────

def coverage_check():
    """(missing, unknown): schema tables no question reaches / wiring typos."""
    from lz_spec import schema
    valid = set()
    for sh in schema.SHEETS:
        if sh.name in schema.INFO_SHEETS or sh.name == "_meta":
            continue
        for t in sh.tables:
            valid.add(f"{sh.name}.{t.name}")

    covered, unknown = set(), []
    for item in Q + APPENDICES:
        for w in item.get("wiring", []):
            table = ".".join(w.split(".")[:2])
            if table not in valid:
                unknown.append(w)
            covered.add(table)

    missing = sorted(t for t in valid if t not in covered and t not in ENGINEER_ONLY)
    stale = sorted(t for t in ENGINEER_ONLY if t not in valid)
    return missing, sorted(set(unknown)) + [f"stale exemption: {s}" for s in stale]


# ── Workbook emit ───────────────────────────────────────────────────────────

def _refs():
    """Assign C1../D1.. refs in catalogue order."""
    c = d = 0
    for q in Q:
        if q["tier"] == "core":
            c += 1
            q["ref"] = f"C{c}"
        else:
            d += 1
            q["ref"] = f"D{d}"


def write_workbook(out: Path):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Theme aligned with lz_spec/gen_template.py (the LLD spec workbook)
    DARK = "1F4E79"     # title band (= template TITLE_FILL)
    HDR = "DDEBF7"      # header row (= template HEADER_FILL)
    BAND = "C6E0B4"     # category band (= template section-band green)
    FILL_IN = "FFFCE5"  # response cells (= template VALUE_FILL)
    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    def _title(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=text)
        c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 26

    def _headers(ws, row, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(name="Calibri", bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=HDR)
            c.alignment = wrap
            c.border = border
        ws.row_dimensions[row].height = 20

    def _est_height(texts_widths, minimum=30):
        lines = 1
        for text, width in texts_widths:
            lines = max(lines, -(-len(text) // width) + text.count("\n"))
        return max(minimum, 14 * lines + 6)

    wb = openpyxl.Workbook()

    # Intro ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Start Here"
    ws.column_dimensions["A"].width = 110
    ws.sheet_view.showGridLines = False
    _title(ws, "Huawei Cloud Landing Zone - Assessment Questionnaire", 1)
    intro = [
        ("", None),
        ("Before you start", "h"),
        ("This questionnaire helps us understand how you want Huawei Cloud to be "
         "organized, secured, connected, and operated. You do not need to have every "
         "answer today. We use your responses to prepare a first design, then review "
         "assumptions and open decisions with you.", None),
        ("", None),
        ("How to complete it", "h"),
        ("- Write as much or as little as you know. If you already have diagrams, "
         "inventories, standards, or policies, attach them instead of retyping the "
         "same information.", None),
        ("- “Unknown”, “not decided yet”, and “no specific requirement” are all "
         "valid answers. Please complete the Core Questions and use the Deep-Dive "
         "Questions only where they apply.", None),
        ("- If a Deep-Dive question does not apply or is left blank, we will propose "
         "a sensible starting point and confirm it with you before implementation.", None),
        ("- Appendices A to C are optional. Use them when accounts, IP ranges, or "
         "team access are easier to describe in a table.", None),
        ("", None),
        ("What happens after this", "h"),
        ("1. We review your answers and prepare a draft landing zone design.", None),
        ("2. We highlight anything that is unclear, undecided, or needs your approval.", None),
        ("3. We review those items with you and finalize the detailed design.", None),
        ("4. The agreed design becomes the baseline for implementation and automation.", None),
        ("", None),
        (f"Version {QUESTIONNAIRE_VERSION} - generated {datetime.date.today():%Y-%m-%d}"
         f" - schema {_schema_version()}", "s"),
    ]
    r = 2
    for text, kind in intro:
        c = ws.cell(row=r, column=1, value=text or None)
        if kind == "h":
            c.font = Font(bold=True, size=11, color=DARK)
        elif kind == "s":
            c.font = Font(italic=True, size=9, color="595959")
        c.alignment = wrap
        if text and kind is None:
            ws.row_dimensions[r].height = _est_height([(text, 105)], minimum=15)
        r += 1

    # Survey sheets ----------------------------------------------------------
    widths = {"A": 7, "B": 46, "C": 40, "D": 34, "E": 50}
    for tier, sheet_name, subtitle in [
        ("core", "Core Questions", "Core Questions"),
        ("deep", "Deep-Dive Questions", "Deep-Dive Questions - answer where relevant"),
    ]:
        ws = wb.create_sheet(sheet_name)
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        _title(ws, f"Huawei Cloud Landing Zone Assessment - {subtitle}", 4)
        _headers(ws, 2, ["No.", "Question", "Guidance", "Example Response", "Customer Response"])
        ws.freeze_panes = "A3"
        r = 3
        for cat in CATEGORIES:
            qs = [q for q in Q if q["tier"] == tier and q["category"] == cat]
            if not qs:
                continue
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            c = ws.cell(row=r, column=1, value=cat)
            c.font = Font(name="Calibri", bold=True, color="375623")
            c.fill = PatternFill("solid", fgColor=BAND)
            c.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[r].height = 18
            r += 1
            for q in qs:
                ex = q["example"] or q["default"]
                cells = [q["ref"], q["question"], q["guidance"], ex or None, None]
                for i, v in enumerate(cells, 1):
                    c = ws.cell(row=r, column=i, value=v)
                    c.alignment = wrap
                    c.border = border
                    if i == 4:
                        c.font = Font(italic=True, size=10, color="595959")
                    if i == 5:
                        c.fill = PatternFill("solid", fgColor=FILL_IN)
                ws.row_dimensions[r].height = _est_height(
                    [(q["question"], 42), (q["guidance"], 36), (ex, 30)])
                r += 1

    # Appendix sheets --------------------------------------------------------
    for ap in APPENDICES:
        ws = wb.create_sheet(ap["name"])
        n = len(ap["columns"])
        for i in range(1, n + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30
        _title(ws, ap["title"], n)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        c = ws.cell(row=2, column=1, value=ap["note"])
        c.font = Font(italic=True, size=9, color="595959")
        c.alignment = wrap
        ws.row_dimensions[2].height = 28
        _headers(ws, 3, ap["columns"])
        ws.freeze_panes = "A4"
        for i, v in enumerate(ap["example"], 1):
            c = ws.cell(row=4, column=i, value=(f"(example) {v}" if i == 1 else v) or None)
            c.font = Font(italic=True, color="595959")
            c.alignment = wrap
            c.border = border
        for r in range(5, 25):
            for i in range(1, n + 1):
                c = ws.cell(row=r, column=i)
                c.alignment = wrap
                c.border = border

    # Hidden wiring sheet ----------------------------------------------------
    ws = wb.create_sheet("_wiring")
    ws.append(["Ref", "Tier", "Category", "Targets", "DefaultIfSilent"])
    for q in Q:
        ws.append([q["ref"], q["tier"], q["category"],
                   "; ".join(q["wiring"]), q["default"]])
    for ap in APPENDICES:
        ws.append([ap["ref"], "appendix", ap["title"],
                   "; ".join(ap["wiring"]), ""])
    ws.append(["_meta", "meta", "",
               f"questionnaire_version={QUESTIONNAIRE_VERSION}; schema_version={_schema_version()}", ""])
    ws.sheet_state = "hidden"

    wb.save(out)


def _schema_version():
    from lz_spec import schema
    return schema.SCHEMA_VERSION


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--out", default=str(DEFAULT_OUT))
    ap.add_argument("--check", action="store_true", help="coverage check only, write nothing")
    args = ap.parse_args(argv)

    _refs()
    missing, unknown = coverage_check()
    for m in missing:
        print(f"ERROR: schema table not covered by any question: {m}", file=sys.stderr)
    for u in unknown:
        print(f"ERROR: wiring target not in schema: {u}", file=sys.stderr)
    if missing or unknown:
        return 1
    core = sum(1 for q in Q if q["tier"] == "core")
    print(f"coverage OK: {core} core + {len(Q) - core} deep questions, "
          f"{len(APPENDICES)} appendices, {len(ENGINEER_ONLY)} exempt tables")
    if args.check:
        return 0
    write_workbook(Path(args.out))
    print(f"wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
