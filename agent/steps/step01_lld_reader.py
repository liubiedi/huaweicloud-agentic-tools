"""Parse an Excel LLD workbook (or JSON dict) into structured lld_json."""
from __future__ import annotations

import json
import os
from typing import Any

import openpyxl
import structlog

log = structlog.get_logger()

# Column indices (0-based) inside each domain sheet
COL_FIELD = 1    # B: field name / terraform variable
COL_VALUE = 2    # C: user-filled value
COL_REQUIRED = 3 # D: "YES" / "NO"
COL_DEFAULT = 4  # E: default value

# Column indices for the Metadata sheet
META_COL_KEY = 0   # A: label
META_COL_VALUE = 1 # B: value

METADATA_KEY_MAP = {
    "Project Name": "project_name",
    "Customer / Organization": "customer",
    "Requested By": "requested_by",
    "Cloud Architect": "cloud_architect",
    "Target Region": "target_region",
    "Target Environment": "target_environment",
    "Requested Deploy Date": "requested_deploy_date",
    "Run ID": "run_id",
}

DOMAIN_SHEETS = ["D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9"]


def _coerce(value: Any, default: Any) -> Any:
    """Return value if non-empty, else default."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return default
    # Normalise boolean strings
    if isinstance(value, str):
        lv = value.strip().lower()
        if lv in ("true", "yes", "1"):
            return True
        if lv in ("false", "no", "0"):
            return False
    return value


def parse_excel(path: str) -> dict[str, Any]:
    """Load an Excel LLD workbook and return structured domain dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, Any] = {}

    # Metadata sheet
    metadata: dict[str, Any] = {}
    if "Metadata" in wb.sheetnames:
        ws = wb["Metadata"]
        for row in ws.iter_rows(values_only=True):
            key_cell = row[META_COL_KEY] if len(row) > META_COL_KEY else None
            val_cell = row[META_COL_VALUE] if len(row) > META_COL_VALUE else None
            if key_cell and str(key_cell) in METADATA_KEY_MAP:
                metadata[METADATA_KEY_MAP[str(key_cell)]] = val_cell
    result["_metadata"] = metadata

    # Domain sheets
    for sheet_name in wb.sheetnames:
        # Match "D1 - Org & Accounts" → "D1"
        domain_id = None
        for d in DOMAIN_SHEETS:
            if sheet_name.startswith(d):
                domain_id = d
                break
        if domain_id is None:
            continue

        ws = wb[sheet_name]
        domain_data: dict[str, Any] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            field = row[COL_FIELD] if len(row) > COL_FIELD else None
            if not field or str(field).startswith("#"):
                continue
            raw_value = row[COL_VALUE] if len(row) > COL_VALUE else None
            default = row[COL_DEFAULT] if len(row) > COL_DEFAULT else None
            domain_data[str(field)] = _coerce(raw_value, _coerce(default, None))

        result[domain_id] = domain_data
        log.info("domain_parsed", domain=domain_id, field_count=len(domain_data))

    return result


def parse_json(data: dict[str, Any]) -> dict[str, Any]:
    """Accept a pre-parsed JSON dict directly (for testing / non-Excel input)."""
    return data


def detect_filled_domains(lld_json: dict[str, Any]) -> list[str]:
    """
    Return domain IDs where at least one field has a non-null value.
    Used to suggest which modules are ready to deploy.
    """
    filled = []
    for d in DOMAIN_SHEETS:
        domain = lld_json.get(d, {})
        if any(v is not None for v in domain.values()):
            filled.append(d)
    return filled


def download_from_obs(obs_key: str) -> str:
    """Download LLD file from OBS to /tmp and return local path."""
    import boto3
    region = os.environ["TF_STATE_REGION"]
    bucket = os.environ["TF_STATE_BUCKET"]
    client = boto3.client(
        "s3",
        endpoint_url=f"https://obs.{region}.myhuaweicloud.com",
        aws_access_key_id=os.environ["HWC_ACCESS_KEY"],
        aws_secret_access_key=os.environ["HWC_SECRET_KEY"],
        region_name=region,
    )
    local = f"/tmp/lld_{obs_key.replace('/', '_')}"
    client.download_file(bucket, obs_key, local)
    return local


async def execute(state, _context) -> None:
    """Step 1: read LLD from OBS key in state.lld_source_key."""
    log.info("step01_start", source_key=state.lld_source_key)

    if state.lld_source_key.endswith(".json"):
        import boto3, json as _json
        region = os.environ["TF_STATE_REGION"]
        bucket = os.environ["TF_STATE_BUCKET"]
        client = boto3.client(
            "s3",
            endpoint_url=f"https://obs.{region}.myhuaweicloud.com",
            aws_access_key_id=os.environ["HWC_ACCESS_KEY"],
            aws_secret_access_key=os.environ["HWC_SECRET_KEY"],
            region_name=region,
        )
        body = client.get_object(Bucket=bucket, Key=state.lld_source_key)["Body"].read()
        state.lld_json = parse_json(_json.loads(body))
    else:
        local_path = download_from_obs(state.lld_source_key)
        state.lld_json = parse_excel(local_path)

    state.lld_metadata = state.lld_json.get("_metadata", {})

    filled = detect_filled_domains(state.lld_json)
    log.info("lld_read_complete", filled_domains=filled)
